# requirements: streamlit, pandas, openpyxl, pytesseract, pillow, opencv-python

import streamlit as st
import pandas as pd
import pytesseract
from PIL import Image
import tempfile
import os
import cv2
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Excel SpeedTest Autofiller", layout="centered")
st.title("📊 Excel SpeedTest Autofiller")
st.write("Upload your Excel file with embedded screenshots. The app will extract speedtest values from images and autofill the corresponding tables.")

def extract_images_from_excel(file):
    wb = load_workbook(file)
    image_data = []
    for sheet in wb.worksheets:
        if not hasattr(sheet, '_images'):
            continue
        for img in sheet._images:
            img_ref = img.ref
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
               def extract_images_from_excel(file_path):
    # ...code to open Excel and locate images...
    for img in images:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
            img.save(tmp_img.name)  # ✅ fixed line
            image_paths.append(tmp_img.name)

                image_data.append((sheet.title, img_ref, tmp_img.name))
    return image_data

def extract_text_from_image(img_path):
    image = cv2.imread(img_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, thresh = cv2.threshold(blur, 150, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    processed_img = cv2.bitwise_not(thresh)
    text = pytesseract.image_to_string(processed_img)
    return text

def parse_speedtest_text(text):
    result = {}
    lines = text.split("\n")
    for line in lines:
        line = line.strip()
        if "Download" in line or "DL" in line:
            result["Download"] = ''.join(filter(str.isdigit, line))
        elif "Upload" in line or "UL" in line:
            result["Upload"] = ''.join(filter(str.isdigit, line))
        elif "Ping" in line:
            result["Ping"] = ''.join(filter(str.isdigit, line))
        elif "Jitter" in line:
            result["Jitter"] = ''.join(filter(str.isdigit, line))
    return result

def autofill_excel(file, parsed_results):
    wb = load_workbook(file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i, result in enumerate(parsed_results):
            row = 10 + i  # Example row logic
            ws[f"C{row}"] = result.get("Download", "")
            ws[f"D{row}"] = result.get("Upload", "")
            ws[f"E{row}"] = result.get("Ping", "")
            ws[f"F{row}"] = result.get("Jitter", "")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

uploaded_file = st.file_uploader("Upload Excel File", type="xlsx")
if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    st.info("Extracting images and processing with OCR...")
    image_paths = extract_images_from_excel(tmp_path)
    results = []
    for sheet, ref, img_path in image_paths:
        text = extract_text_from_image(img_path)
        parsed = parse_speedtest_text(text)
        results.append(parsed)

    st.success(f"Extracted and parsed {len(results)} screenshot(s)")
    output_excel = autofill_excel(tmp_path, results)
    st.download_button("Download Completed Excel", data=output_excel, file_name="Autofilled_Result.xlsx")
